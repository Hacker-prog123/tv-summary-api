def run_tv_summary(main_file_bytes, rr_file_bytes):
    import pandas as pd
    from openpyxl import load_workbook
    from io import BytesIO

    main_file = BytesIO(main_file_bytes)
    rr_file = BytesIO(rr_file_bytes)

    df = pd.read_excel(main_file, sheet_name="TV")
    esn_df = pd.read_excel(main_file, sheet_name="ESN", usecols="B:E")
    rr_df_full = pd.read_excel(rr_file, sheet_name="Export")

    # ✅ Fix for key errors on column names
    df.columns = df.columns.str.strip()
    rr_df_full.columns = rr_df_full.columns.str.strip()

    def clean_esn(val):
        val = str(val).strip().replace("\n", "").replace("\r", "").replace("\xa0", "").replace(" ", "")
        if val.replace('.', '', 1).isdigit():
            val = str(int(float(val)))
        return val

    df["Engine Number (Ex)"] = df["Engine Number (Ex)"].map(clean_esn)
    current_week_esns = pd.concat([esn_df[col] for col in esn_df.columns]).dropna().map(clean_esn).unique()
    df_current_week = df[df["Engine Number (Ex)"].isin(current_week_esns)].copy()

    def classify_delivery_type(row):
        status = str(row["SAESL Status"]).strip()
        gate = row.get("Gate", None)
        tvr_gate = str(row.get("TVR Gate Raised", "")).strip().upper()
        if pd.isna(status) or pd.isna(gate): return "Undelivered"
        try: gate = int(gate)
        except: return "Undelivered"
        if gate == 1 and status == "Under Review (approved TV)": return "TV Delivered"
        if gate == 1 and tvr_gate == "0": return "Undelivered"
        if gate == 1 and tvr_gate == "1" and status == "Under Review (approved TV)": return "TV Delivered"
        if gate == 1 and status == "Cleared" and tvr_gate == "G0": return "TV Delivered"
        if gate == 2 and status == "Cleared" and tvr_gate in ["G0", "G1"]: return "TV Delivered"
        if gate == 2 and status == "Under Review (approved TV)" and tvr_gate in ["G0", "G1"]: return "TV Delivered"
        if gate in [1, 2] and status in ["Under Review (draft)", "Under Review (MP Release)", "Cleared", "Cleared Mitigated"]:
            return "HU/Draft Delivered"
        if gate in [3, 4] and status in ["Cleared", "Cleared Mitigated", "Under Review (approved TV)"]:
            return "TV Delivered"
        if status in ["Need HU", "Need Draft", "Draft Clarification", "Need Approved TV"]:
            return "Undelivered"
        return "Undelivered"

    df["Delivery Type"] = df.apply(classify_delivery_type, axis=1)
    df_current_week["ESN"] = df_current_week["Engine Number (Ex)"]

    current_week_summary = pd.DataFrame({
        "Metric": [
            "Total ESNs in Current Week", "TV Delivered", "HU/Draft Delivered", "Undelivered", "Delivery Rate (%)"
        ],
        "Value": [
            len(df_current_week),
            (df_current_week["Delivery Type"] == "TV Delivered").sum(),
            (df_current_week["Delivery Type"] == "HU/Draft Delivered").sum(),
            (df_current_week["Delivery Type"] == "Undelivered").sum(),
            round(df_current_week["Delivery Type"].isin(["TV Delivered", "HU/Draft Delivered"]).sum()/len(df_current_week)*100, 2) if len(df_current_week) else 0
        ]
    })

    current_week_detail = df_current_week.groupby(["ESN", "Gate"]).agg(
        HU_Draft_Delivered=("Delivery Type", lambda x: (x == "HU/Draft Delivered").sum()),
        TV_Delivered=("Delivery Type", lambda x: (x == "TV Delivered").sum()),
        Undelivered=("Delivery Type", lambda x: (x == "Undelivered").sum())
    ).reset_index()
    current_week_detail["Total Delivered"] = current_week_detail["HU_Draft_Delivered"] + current_week_detail["TV_Delivered"]
    current_week_detail["Total Count"] = current_week_detail["Total Delivered"] + current_week_detail["Undelivered"]
    current_week_detail["Delivery %"] = round(current_week_detail["Total Delivered"] / current_week_detail["Total Count"] * 100, 2)

    delivery_type_breakdown = df["Delivery Type"].value_counts().reset_index()
    delivery_type_breakdown.columns = ["Delivery Type", "Count"]

    gate_perf_all = df.groupby("Gate").agg(
        Total_TVs=("Custom ID", "count"),
        Delivered=("Delivery Type", lambda x: x.isin(["TV Delivered", "HU/Draft Delivered"]).sum()),
        TV_Delivered=("Delivery Type", lambda x: (x == "TV Delivered").sum()),
        HU_Draft_Delivered=("Delivery Type", lambda x: (x == "HU/Draft Delivered").sum()),
        Undelivered=("Delivery Type", lambda x: (x == "Undelivered").sum())
    ).reset_index()
    gate_perf_all["Delivery %"] = round(gate_perf_all["Delivered"] / gate_perf_all["Total_TVs"] * 100, 2)

    summary_all_df = pd.DataFrame({
        "Metric": ["Total TVs in List", "Total Delivered (TV or HU)", "TV Delivered", "HU/Draft Delivered", "Total Undelivered", "Delivery Rate (%)"],
        "Value": [
            len(df),
            df["Delivery Type"].isin(["TV Delivered", "HU/Draft Delivered"]).sum(),
            (df["Delivery Type"] == "TV Delivered").sum(),
            (df["Delivery Type"] == "HU/Draft Delivered").sum(),
            (df["Delivery Type"] == "Undelivered").sum(),
            round(df["Delivery Type"].isin(["TV Delivered", "HU/Draft Delivered"]).sum()/len(df)*100, 2) if len(df) else 0
        ]
    })

    rr_file.seek(0)
    wb = load_workbook(filename=rr_file, data_only=True)
    ws = wb["Export"]
    green_tv_nos = [str(row[0].value).strip() for row in ws.iter_rows(min_row=2) if row[0].fill.start_color.rgb in ("FF92D050")]

    rr_df = rr_df_full[
        (rr_df_full["Tag delivered?"] == "Yes") &
        (rr_df_full["Applicant"].str.strip() == "SAESL") &
        (rr_df_full["TV NO"].astype(str).str.strip().isin(green_tv_nos))
    ].copy()
    rr_df["Custom ID"] = rr_df["TV NO"].astype(str).str.strip()
    df["Custom ID"] = df["Custom ID"].astype(str).str.strip()
    rr_df["Your Status"] = rr_df["Custom ID"].map(df.set_index("Custom ID")["Delivery Type"].to_dict()).fillna("Not Found")
    rr_df["Actual SAESL Status"] = rr_df["Custom ID"].map(df.set_index("Custom ID")["SAESL Status"].to_dict())

    rr_discrepancies = rr_df[rr_df["Your Status"] == "Undelivered"][
        ["Custom ID", "Gate", "Your Status", "Actual SAESL Status", "Tag delivered?"]
    ].copy()
    rr_discrepancies.columns = ["TV Number", "Gate", "Your Status", "Actual SAESL Status", "RR Tag Delivered"]

    rr_gate_summary = rr_df.groupby("Gate").agg(RR_Tagged_Delivered=("Custom ID", "count")).reset_index()
    true_status_breakdown = rr_df.groupby(["Gate", "Your Status"]).agg(Count=("Custom ID", "count")).reset_index()
    true_status_pivot = true_status_breakdown.pivot(index="Gate", columns="Your Status", values="Count").fillna(0).reset_index()
    true_status_pivot.columns.name = None

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_all_df.to_excel(writer, sheet_name="Summary (All Tags)", index=False)
        current_week_summary.to_excel(writer, sheet_name="Current Week Summary", index=False)
        current_week_detail.to_excel(writer, sheet_name="Current Week ESN Detail", index=False)
        delivery_type_breakdown.to_excel(writer, sheet_name="Delivery Breakdown", index=False)
        gate_perf_all.to_excel(writer, sheet_name="Gate Perf (All TVs)", index=False)
        rr_discrepancies.to_excel(writer, sheet_name="RR Discrepancy", index=False)
        rr_gate_summary.to_excel(writer, sheet_name="RR Gate Summary", index=False)
        true_status_pivot.to_excel(writer, sheet_name="RR True Delivery Breakdown", index=False)
        df.to_excel(writer, sheet_name="Raw Data", index=False)

    output.seek(0)
    return output
